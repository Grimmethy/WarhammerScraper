import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_SHOP_BASE     = "https://www.warhammer.com/en-US/shop/"
_IMG_BASE      = "https://www.warhammer.com/app/resources/catalog/product/920x950/"
_THUMB_W       = 100
_THUMB_H       = 100
_ROW_HEIGHT_PT = 80
_HEADER_BG     = "1F3864"
_HEADER_FG     = "FFFFFF"
_LINK_COLOR    = "0563C1"
_COLUMNS       = [("#", 5), ("Item Name", 42), ("Price", 12), ("Website Link", 55), ("Image", 16), ("Image URL", 60)]


class WorksheetBuilder:
    def write(
        self,
        title: str,
        output_path: str,
        products: list[tuple],
        manifest: dict[str, io.BytesIO | None],
    ) -> None:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        n = len(products)
        print(f"\nBuilding {title} -- {n} products")

        wb = Workbook()
        ws = wb.active
        ws.title = title

        hdr_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        hdr_font = Font(color=_HEADER_FG, bold=True, size=11)
        center   = Alignment(horizontal="center", vertical="center")
        mid_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for ci, (label, width) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 20

        for i, (name, price, slug, img_file) in enumerate(products, 1):
            row = i + 1
            ws.row_dimensions[row].height = _ROW_HEIGHT_PT

            ws.cell(row=row, column=1, value=i).alignment = center
            ws.cell(row=row, column=2, value=name).alignment = mid_left
            ws.cell(row=row, column=3, value=price).alignment = center

            url = _SHOP_BASE + slug
            lc  = ws.cell(row=row, column=4, value=url)
            lc.hyperlink = url
            lc.font      = Font(color=_LINK_COLOR, underline="single")
            lc.alignment = mid_left

            buf = manifest.get(img_file)
            if buf:
                xl_img        = XLImage(buf)
                xl_img.width  = _THUMB_W
                xl_img.height = _THUMB_H
                ws.add_image(xl_img, f"E{row}")
            else:
                print(f"  [warn] image missing: {img_file}")

            ws.cell(row=row, column=6, value=_IMG_BASE + img_file).alignment = mid_left
            print(f"  [{i:02d}/{n}] {name}")

        wb.save(output_path)
        print(f"Done: {output_path}")
