#!/usr/bin/env python3
"""
Warhammer Category Scraper
Scrapes all products from a Warhammer store category page and writes them to
an .xlsx file with Name, Price, URL, embedded thumbnail, and image URL.

Setup (once):
    pip install -r requirements.txt
    playwright install firefox

Usage:
    python scraper.py --url "https://www.warhammer.com/en-US/shop/age-of-sigmar/grand-alliance-destruction/gloomspite-gitz" --output gloomspite_gitz.xlsx
    python scraper.py --url "..." --output out.xlsx --images ./img/faction
"""

import argparse
import asyncio
import io
import re
import sys
from pathlib import Path
from urllib.parse import urljoin, urlparse, urlunparse

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

BASE_URL         = "https://www.warhammer.com"
THUMB_W          = 100          # thumbnail width in xlsx (pixels)
THUMB_H          = 100          # thumbnail height in xlsx (pixels)
ROW_HEIGHT_PT    = 80           # xlsx row height in points
IMG_TIMEOUT      = 20           # seconds per image download
PAGE_WAIT_MS     = 2500         # ms to wait after each Show More click

HEADER_BG = "1F3864"
HEADER_FG = "FFFFFF"
LINK_COLOR = "0563C1"

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64; rv:125.0) Gecko/20100101 Firefox/125.0"
    )
}

# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

def clean_url(href: str) -> str:
    """Strip tracking query params and make absolute."""
    if not href:
        return ""
    full = urljoin(BASE_URL, href)
    parsed = urlparse(full)
    # Drop queryID and similar tracking params
    return urlunparse(parsed._replace(query="", fragment=""))


async def dismiss_overlays(page) -> None:
    for sel in [
        "#onetrust-accept-btn-handler",
        'button:has-text("Allow All")',
        'button:has-text("Accept All")',
        'button:has-text("Accept Cookies")',
    ]:
        try:
            await page.click(sel, timeout=3000)
            await page.wait_for_timeout(400)
            return
        except PWTimeout:
            pass


async def extract_products(page) -> list[dict]:
    """Pull every visible product card off the current page state."""
    return await page.evaluate("""() => {
        const BASE = "https://www.warhammer.com";
        const cards = document.querySelectorAll('[data-testid="product-list-item"]');
        const results = [];

        for (const card of cards) {
            const nameEl  = card.querySelector('[data-testid="product-card-name"]');
            const priceEl = card.querySelector('[data-testid="product-card-current-price"]');
            const linkEl  = card.querySelector('[data-testid="product-card-details"]');
            const imgEl   = card.querySelector('[data-testid="product-card-image"] img');

            const name  = nameEl  ? nameEl.textContent.trim()  : "";
            const price = priceEl ? priceEl.textContent.trim() : "";

            // Clean the product URL — strip queryID tracking param
            let link = "";
            if (linkEl) {
                const href = linkEl.getAttribute("href") || "";
                const url = new URL(href, BASE);
                url.search = "";
                link = url.toString();
            }

            // Image: src is a relative path like /app/resources/catalog/product/920x950/...
            let imageUrl = "";
            if (imgEl) {
                const src = imgEl.getAttribute("src") || "";
                imageUrl = src.startsWith("http") ? src : BASE + src;
            }

            if (name && link) {
                results.push({ name, price, link, imageUrl });
            }
        }
        return results;
    }""")


async def scrape_category(url: str) -> list[dict]:
    all_products: list[dict] = []
    seen_links: set[str] = set()

    async with async_playwright() as pw:
        browser = await pw.firefox.launch(headless=True)
        ctx = await browser.new_context(
            user_agent=HTTP_HEADERS["User-Agent"],
            viewport={"width": 1440, "height": 900},
        )
        page = await ctx.new_page()

        print(f"  Loading: {url}")
        await page.goto(url, wait_until="domcontentloaded", timeout=60_000)
        await dismiss_overlays(page)
        await page.wait_for_timeout(PAGE_WAIT_MS)

        batch_num = 1
        while True:
            products = await extract_products(page)
            new = [p for p in products if p["link"] not in seen_links]
            for p in new:
                seen_links.add(p["link"])
            all_products.extend(new)
            print(f"  Batch {batch_num}: {len(new)} new products (total: {len(all_products)})")

            # Click "Show More" if present and visible
            show_more = await page.query_selector('[data-testid="button-show-more"]')
            if show_more and await show_more.is_visible():
                await show_more.scroll_into_view_if_needed()
                await show_more.click()
                try:
                    await page.wait_for_function(
                        f"document.querySelectorAll('[data-testid=\"product-list-item\"]').length > {len(all_products)}",
                        timeout=15_000,
                    )
                except PWTimeout:
                    await page.wait_for_timeout(PAGE_WAIT_MS)
                batch_num += 1
            else:
                break

        await browser.close()

    return all_products


# ---------------------------------------------------------------------------
# Images
# ---------------------------------------------------------------------------

def download_image(url: str, dest_dir: Path) -> Path | None:
    if not url:
        return None
    try:
        resp = requests.get(url, headers=HTTP_HEADERS, timeout=IMG_TIMEOUT)
        resp.raise_for_status()
        ct = resp.headers.get("content-type", "")
        ext = ".png" if "png" in ct else ".webp" if "webp" in ct else ".jpg"
        # Name from the URL path tail, sanitised
        tail = url.rstrip("/").split("/")[-1].split("?")[0]
        safe = re.sub(r"[^\w.-]", "_", tail)[:80]
        path = dest_dir / (safe + ext)
        path.write_bytes(resp.content)
        return path
    except Exception as e:
        print(f"    [warn] image download failed ({url[:60]}): {e}")
        return None


def make_thumbnail(src_path: Path) -> io.BytesIO | None:
    try:
        img = PILImage.open(src_path).convert("RGBA")
        img.thumbnail((THUMB_W, THUMB_H), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"    [warn] thumbnail failed for {src_path.name}: {e}")
        return None


# ---------------------------------------------------------------------------
# xlsx output
# ---------------------------------------------------------------------------

COLUMNS = [
    ("#",            5),
    ("Item Name",   42),
    ("Price",       12),
    ("Website Link", 55),
    ("Image",       16),
    ("Image URL",   60),
]


def write_xlsx(products: list[dict], image_dir: Path, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    hdr_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    hdr_font = Font(color=HEADER_FG, bold=True, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    mid_left = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for col_idx, (label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    print(f"\nDownloading images and writing rows...")
    for i, product in enumerate(products, 1):
        row = i + 1
        ws.row_dimensions[row].height = ROW_HEIGHT_PT

        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=product["name"]).alignment = mid_left
        ws.cell(row=row, column=3, value=product["price"]).alignment = center

        link_cell = ws.cell(row=row, column=4, value=product["link"])
        link_cell.hyperlink = product["link"]
        link_cell.font = Font(color=LINK_COLOR, underline="single")
        link_cell.alignment = mid_left

        # Download image and embed thumbnail
        img_url = product.get("imageUrl", "")
        if img_url:
            local = download_image(img_url, image_dir)
            if local:
                buf = make_thumbnail(local)
                if buf:
                    xl_img = XLImage(buf)
                    xl_img.width  = THUMB_W
                    xl_img.height = THUMB_H
                    ws.add_image(xl_img, f"E{row}")

        ws.cell(row=row, column=6, value=img_url).alignment = mid_left

        if i % 10 == 0 or i == len(products):
            print(f"  {i}/{len(products)} rows complete")

    wb.save(output_path)
    print(f"\nSaved → {output_path}")
    print(f"Images → {image_dir.resolve()}/")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

async def run(url: str, output: str, images: str) -> None:
    image_dir = Path(images)
    image_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n=== Scraping ===")
    products = await scrape_category(url)
    print(f"\nTotal products: {len(products)}")

    if not products:
        print("No products found — double-check the URL and try again.")
        sys.exit(1)

    write_xlsx(products, image_dir, output)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scrape a Warhammer store category page to .xlsx"
    )
    parser.add_argument("--url",    required=True, help="Category page URL")
    parser.add_argument("--output", default="warhammer_products.xlsx", help="Output .xlsx file")
    parser.add_argument("--images", default="images", help="Directory for downloaded images")
    args = parser.parse_args()
    asyncio.run(run(args.url, args.output, args.images))


if __name__ == "__main__":
    main()
