#!/usr/bin/env python3
"""
Generic Warhammer faction xlsx builder.

Usage:
  python build.py <faction_key> [<faction_key> ...]
  python build.py --all
  python build.py --list

Examples:
  python build.py flesh_eater_courts
  python build.py nighthaunt ossiarch_bonereapers soulblight_gravelords
  python build.py --all
"""

import io
import sys
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from factions import FACTIONS

BASE_URL  = "https://www.warhammer.com"
IMG_BASE  = BASE_URL + "/app/resources/catalog/product/920x950/"
SHOP_BASE = BASE_URL + "/en-US/shop/"
IMG_DIR   = Path("data/images")

THUMB_W        = 100
THUMB_H        = 100
ROW_HEIGHT_PT  = 80
IMG_TIMEOUT    = 20
HEADER_BG      = "1F3864"
HEADER_FG      = "FFFFFF"
LINK_COLOR     = "0563C1"
HTTP_HEADERS   = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:125.0) Gecko/20100101 Firefox/125.0"}
COLUMNS        = [("#", 5), ("Item Name", 42), ("Price", 12), ("Website Link", 55), ("Image", 16), ("Image URL", 60)]


def download_image(filename, dest_dir):
    try:
        r = requests.get(IMG_BASE + filename, headers=HTTP_HEADERS, timeout=IMG_TIMEOUT)
        r.raise_for_status()
        p = dest_dir / filename
        p.write_bytes(r.content)
        return p
    except Exception as e:
        print(f"  [warn] {filename}: {e}")
        return None


def make_thumbnail(src):
    try:
        img = PILImage.open(src).convert("RGBA")
        img.thumbnail((THUMB_W, THUMB_H), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"  [warn] thumbnail {src.name}: {e}")
        return None


def build_faction(key):
    faction = FACTIONS[key]
    title    = faction["title"]
    output   = faction["output"]
    products = faction["products"]
    n        = len(products)

    Path(output).parent.mkdir(parents=True, exist_ok=True)
    IMG_DIR.mkdir(parents=True, exist_ok=True)

    print(f"\nBuilding {title} -- {n} products")

    wb = Workbook()
    ws = wb.active
    ws.title = title

    hdr_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    hdr_font = Font(color=HEADER_FG, bold=True, size=11)
    center   = Alignment(horizontal="center", vertical="center")
    mid_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for ci, (label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 20

    for i, (name, price, slug, img_file) in enumerate(products, 1):
        row = i + 1
        ws.row_dimensions[row].height = ROW_HEIGHT_PT

        ws.cell(row=row, column=1, value=i).alignment = center
        ws.cell(row=row, column=2, value=name).alignment = mid_left
        ws.cell(row=row, column=3, value=price).alignment = center

        url = SHOP_BASE + slug
        lc  = ws.cell(row=row, column=4, value=url)
        lc.hyperlink = url
        lc.font      = Font(color=LINK_COLOR, underline="single")
        lc.alignment = mid_left

        local = download_image(img_file, IMG_DIR)
        if local:
            buf = make_thumbnail(local)
            if buf:
                xl_img        = XLImage(buf)
                xl_img.width  = THUMB_W
                xl_img.height = THUMB_H
                ws.add_image(xl_img, f"E{row}")

        ws.cell(row=row, column=6, value=IMG_BASE + img_file).alignment = mid_left
        print(f"  [{i:02d}/{n}] {name}")

    wb.save(output)
    print(f"Done: {output}")


def main():
    args = sys.argv[1:]

    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    if args[0] == "--list":
        print("Available factions:")
        for key, f in FACTIONS.items():
            print(f"  {key:<30} ({len(f['products'])} products)  ->  {f['output']}")
        sys.exit(0)

    keys = list(FACTIONS.keys()) if args[0] == "--all" else args

    unknown = [k for k in keys if k not in FACTIONS]
    if unknown:
        print(f"Unknown faction(s): {', '.join(unknown)}")
        print("Run 'python build.py --list' to see available keys.")
        sys.exit(1)

    for key in keys:
        build_faction(key)

    print(f"\nAll done. Built {len(keys)} faction(s).")


if __name__ == "__main__":
    main()
